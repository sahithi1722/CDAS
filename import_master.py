from openpyxl import load_workbook

from masters.models import Shop, Equipment, SubEquipment

wb = load_workbook("master_data.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2, values_only=True):
    shop_code, shop_desc, eqpt_code, sub_code = row

    shop, _ = Shop.objects.get_or_create(
        shop_code=str(shop_code).zfill(2),
        defaults={"shop_desc": shop_desc},
    )

    equipment, _ = Equipment.objects.get_or_create(
        shop=shop,
        eqpt_code=str(eqpt_code),
        defaults={"eqpt_desc": str(eqpt_code)},
    )

    if sub_code and str(sub_code).strip():
        SubEquipment.objects.get_or_create(
            equipment=equipment,
            sub_eqpt_code=str(sub_code),
            defaults={"sub_eqpt_desc": str(sub_code)},
        )

print("Master data imported successfully!")
print("Shops:", Shop.objects.count())
print("Equipment:", Equipment.objects.count())
print("SubEquipment:", SubEquipment.objects.count())